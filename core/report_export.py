"""
Report export module for Af Analyzer.

Handles generation of static analysis figures using Matplotlib for export,
and Excel report generation.
"""

import io
from datetime import datetime
from typing import Optional, Tuple

import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Use a font that supports Chinese characters
matplotlib.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

# Chart colors from design-system/af-analyzer/MASTER.md (Morandi Palette)
CHART_COLORS = {
    "primary": "#8B9DC3",      # Morandi Blue - main curve
    "highlight": "#E8E4E0",    # Light white
    "grid": "rgba(255,255,255,0.1)",
    "axis_text": "#8A8A9A",
    "text": "#E8E4E0",
    "bg_deep": "#0F0F1A",
    "bg_surface": "#1A1A2E",
}

TANGENT_COLORS = {
    "low_baseline": "#A4B8A4",   # Morandi Green
    "high_baseline": "#D4B896",  # Morandi Apricot
    "middle_tangent": "#C4A4A4", # Morandi Rose
}

MARKER_COLORS = {
    "as_marker": "#A4B8A4",      # Green for As
    "af_marker": "#D4B896",      # Apricot for Af-tan
}


def generate_analysis_figure(
    temps: np.ndarray,
    values: np.ndarray,
    As: float,
    Af_tan: float,
    max_slope_temp: float,
    low_baseline: Tuple[float, float],
    high_baseline: Tuple[float, float],
    tangent: Tuple[float, float],
    channel_name: str = "Channel"
) -> plt.Figure:
    """
    Generate a static analysis figure using Matplotlib.

    The figure includes:
    - Main smoothed curve (solid line) - Morandi Blue
    - Three tangent lines (dashed) - Morandi colors
    - As and Af-tan intersection points with annotations

    Args:
        temps: Temperature values (1D array)
        values: Smoothed displacement values (1D array)
        As: Austenite start temperature
        Af_tan: Austenite finish temperature (tangent method)
        max_slope_temp: Temperature at maximum slope point
        low_baseline: (slope, intercept) of low-temperature baseline
        high_baseline: (slope, intercept) of high-temperature baseline
        tangent: (slope, intercept) of middle tangent line
        channel_name: Channel name for title

    Returns:
        Matplotlib Figure object
    """
    # Create figure with higher DPI for export
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)

    temp_min, temp_max = temps.min(), temps.max()
    temp_margin = (temp_max - temp_min) * 0.05

    # 1. Plot main curve - Morandi Blue
    ax.plot(temps, values, color=CHART_COLORS["primary"], linewidth=2.5, label='平滑曲线', zorder=3)

    # 2. Plot tangent lines
    # Low-temperature baseline - Morandi Green, dashed
    low_slope, low_intercept = low_baseline
    low_x = np.array([temp_min, As if not np.isnan(As) else temp_min + temp_margin])
    low_y = low_slope * low_x + low_intercept
    ax.plot(low_x, low_y, color=TANGENT_COLORS["low_baseline"], linestyle='--', linewidth=1.5, label='低温基准线', zorder=2)

    # High-temperature baseline - Morandi Apricot, dashed
    high_slope, high_intercept = high_baseline
    high_x = np.array([Af_tan if not np.isnan(Af_tan) else temp_max - temp_margin, temp_max])
    high_y = high_slope * high_x + high_intercept
    ax.plot(high_x, high_y, color=TANGENT_COLORS["high_baseline"], linestyle='--', linewidth=1.5, label='高温基准线', zorder=2)

    # Middle tangent line - Morandi Rose, dashed
    tangent_slope, tangent_intercept = tangent
    tangent_x = np.array([temp_min, temp_max])
    tangent_y = tangent_slope * tangent_x + tangent_intercept
    ax.plot(tangent_x, tangent_y, color=TANGENT_COLORS["middle_tangent"], linestyle='--', linewidth=1.5, label='中间切线', zorder=2)

    # 3. Plot intersection points
    # As point - Morandi Green
    if not np.isnan(As):
        as_y = tangent_slope * As + tangent_intercept
        ax.plot(As, as_y, 'o', color=MARKER_COLORS["as_marker"], markersize=10, markeredgecolor='white', markeredgewidth=1.5, zorder=4)
        ax.annotate(f'As = {As:.1f}°C', xy=(As, as_y), xytext=(10, 20), textcoords='offset points',
                   fontsize=11, color=MARKER_COLORS["as_marker"], fontweight='bold',
                   bbox=dict(boxstyle='round,pad=0.5', facecolor='#1A1A2E', edgecolor=MARKER_COLORS["as_marker"], alpha=0.9),
                   arrowprops=dict(arrowstyle='->', color=MARKER_COLORS["as_marker"]))

    # Af-tan point - Morandi Apricot
    if not np.isnan(Af_tan):
        af_y = tangent_slope * Af_tan + tangent_intercept
        ax.plot(Af_tan, af_y, 'o', color=MARKER_COLORS["af_marker"], markersize=10, markeredgecolor='white', markeredgewidth=1.5, zorder=4)
        ax.annotate(f'Af-tan = {Af_tan:.1f}°C', xy=(Af_tan, af_y), xytext=(-10, 20), textcoords='offset points',
                   fontsize=11, color=MARKER_COLORS["af_marker"], fontweight='bold', ha='right',
                   bbox=dict(boxstyle='round,pad=0.5', facecolor='#1A1A2E', edgecolor=MARKER_COLORS["af_marker"], alpha=0.9),
                   arrowprops=dict(arrowstyle='->', color=MARKER_COLORS["af_marker"]))

    # Max slope point - Morandi Rose diamond
    if not np.isnan(max_slope_temp):
        max_y = tangent_slope * max_slope_temp + tangent_intercept
        ax.plot(max_slope_temp, max_y, 'D', color=TANGENT_COLORS["middle_tangent"], markersize=6, zorder=3)

    # Formatting
    ax.set_xlabel('Temperature (°C)', fontsize=12, color=CHART_COLORS["axis_text"])
    ax.set_ylabel('Displacement', fontsize=12, color=CHART_COLORS["axis_text"])
    ax.set_title(f'{channel_name} 切线分析', fontsize=14, fontweight='bold', color=CHART_COLORS["text"], pad=15)
    ax.grid(True, color=(1.0, 1.0, 1.0, 0.1), linestyle='-', linewidth=0.5)
    ax.legend(loc='upper left', fontsize=10, framealpha=0.9, facecolor='#1A1A2E', edgecolor=(1.0, 1.0, 1.0, 0.1), labelcolor=CHART_COLORS["text"])

    # Set tick colors
    ax.tick_params(axis='both', colors=CHART_COLORS["axis_text"], labelsize=10)

    # Set spine colors
    for spine in ax.spines.values():
        spine.set_edgecolor(CHART_COLORS["axis_text"])
        spine.set_linewidth(0.8)

    # Set dark background
    ax.set_facecolor('#0F0F1A')
    fig.patch.set_facecolor('#0F0F1A')

    # Add result annotation at bottom with dark background
    as_valid = not np.isnan(As)
    af_valid = not np.isnan(Af_tan)

    if as_valid and af_valid:
        result_text = f'分析结果: As = {As:.1f}°C, Af-tan = {Af_tan:.1f}°C'
        fig.text(0.5, 0.02, result_text, ha='center', fontsize=11, color=CHART_COLORS["text"],
                 transform=fig.transFigure, bbox=dict(boxstyle='round,pad=0.5', facecolor='#1A1A2E', edgecolor=(1.0, 1.0, 1.0, 0.1)))
    else:
        # Partial result - show warning with available data
        parts = []
        if as_valid:
            parts.append(f'As = {As:.1f}°C')
        else:
            parts.append('As = N/A')
        if af_valid:
            parts.append(f'Af-tan = {Af_tan:.1f}°C')
        else:
            parts.append('Af-tan = N/A')
        result_text = f'[警告] 部分数据缺失 | {", ".join(parts)}'
        fig.text(0.5, 0.02, result_text, ha='center', fontsize=11, color='#FFA500',
                 transform=fig.transFigure, bbox=dict(boxstyle='round,pad=0.5', facecolor='#1A1A2E', edgecolor='#FFA500'))

    plt.tight_layout(rect=[0, 0.05, 1, 1])  # Make room for bottom text
    return fig


def figure_to_png_bytes(fig: plt.Figure, dpi: int = 300) -> bytes:
    """
    Convert Matplotlib figure to PNG bytes.

    Args:
        fig: Matplotlib Figure object
        dpi: Resolution for the output image

    Returns:
        PNG image as bytes
    """
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='#0F0F1A')
    buf.seek(0)
    png_bytes = buf.read()
    buf.close()
    return png_bytes


def export_excel_report(
    file_name: str,
    channel: str,
    result: dict,
    smooth_params: Tuple[int, int],
    low_range: Tuple[float, float],
    high_range: Tuple[float, float],
    slope_offset: int
) -> bytes:
    """
    Generate Excel report with analysis results and processed data.

    The Excel file contains two sheets:
    - Sheet1 "分析结果": Summary of analysis parameters and results
    - Sheet2 "处理数据": Preprocessed temperature-displacement data

    Args:
        file_name: Original data file name
        channel: Channel name analyzed
        result: Analysis result dictionary
        smooth_params: (window_length, polyorder) for smoothing
        low_range: (t_start, t_end) for low-temperature baseline
        high_range: (t_start, t_end) for high-temperature baseline
        slope_offset: Middle tangent offset value

    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14, color="FFFFFF")
    normal_font = Font(size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

    # Sheet 1: 分析结果
    ws1 = wb.create_sheet("分析结果")

    # Title
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "Af Analyzer - 奥氏体相变温度分析报告"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center_align
    ws1['A1'].fill = header_fill

    # Report info
    ws1['A3'] = "报告生成时间"
    ws1['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws1['A3'].font = header_font

    # Check for incomplete analysis
    as_val = result.get('As', np.nan)
    af_val = result.get('Af_tan', np.nan)
    is_incomplete = np.isnan(as_val) or np.isnan(af_val)

    if is_incomplete:
        ws1['A4'] = "分析状态"
        ws1['B4'] = "⚠ 分析不完整（部分结果缺失）"
        ws1['A4'].font = header_font
        ws1['B4'].font = Font(size=11, color="FF8C00")  # Orange warning color

    # Analysis results section
    row = 5
    ws1.cell(row, 1, "分析结果")
    ws1.cell(row, 1).font = title_font
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f'A{row}:D{row}')

    row += 2
    # Format values, showing N/A for NaN
    as_display = f"{result.get('As', np.nan):.1f} °C" if not np.isnan(result.get('As', np.nan)) else "N/A"
    af_display = f"{result.get('Af_tan', np.nan):.1f} °C" if not np.isnan(result.get('Af_tan', np.nan)) else "N/A"
    max_slope_display = f"{result.get('max_slope_temp', np.nan):.1f} °C" if not np.isnan(result.get('max_slope_temp', np.nan)) else "N/A"

    results_data = [
        ("数据文件", file_name),
        ("分析通道", channel),
        ("As (奥氏体开始)", as_display),
        ("Af-tan (奥氏体完成)", af_display),
        ("最大斜率点温度", max_slope_display),
    ]

    for label, value in results_data:
        ws1.cell(row, 1, label)
        ws1.cell(row, 1).font = header_font
        ws1.cell(row, 2, value)
        ws1.cell(row, 2).font = normal_font
        row += 1

    # Analysis parameters section
    row += 1
    ws1.cell(row, 1, "分析参数")
    ws1.cell(row, 1).font = title_font
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f'A{row}:D{row}')

    row += 2
    smooth_window, smooth_polyorder = smooth_params
    params_data = [
        ("平滑窗口大小", f"{smooth_window}"),
        ("多项式阶数", f"{smooth_polyorder}"),
        ("低温基准线区间", f"{low_range[0]:.1f}°C ~ {low_range[1]:.1f}°C"),
        ("高温基准线区间", f"{high_range[0]:.1f}°C ~ {high_range[1]:.1f}°C"),
        ("中间切线偏移", f"{slope_offset:+d}"),
    ]

    for label, value in params_data:
        ws1.cell(row, 1, label)
        ws1.cell(row, 1).font = header_font
        ws1.cell(row, 2, value)
        ws1.cell(row, 2).font = normal_font
        row += 1

    # Apply borders to data cells
    for r in range(3, row):
        for c in [1, 2]:
            ws1.cell(r, c).border = border_thin

    # Adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 30

    # Sheet 2: 处理数据
    ws2 = wb.create_sheet("处理数据")

    # Title
    ws2.merge_cells('A1:C1')
    ws2['A1'] = "处理后的温度-位移数据"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center_align
    ws2['A1'].fill = header_fill

    # Header row
    headers = ["序号", "Temperature (°C)", "Displacement"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(3, col, header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin

    # Data rows
    temps = result.get("temps_smooth", np.array([]))
    values = result.get("values_smooth", np.array([]))

    for i, (t, v) in enumerate(zip(temps, values), 1):
        ws2.cell(3 + i, 1, i)
        ws2.cell(3 + i, 1).font = normal_font
        ws2.cell(3 + i, 1).alignment = center_align
        ws2.cell(3 + i, 1).border = border_thin

        ws2.cell(3 + i, 2, f"{t:.2f}")
        ws2.cell(3 + i, 2).font = normal_font
        ws2.cell(3 + i, 2).alignment = center_align
        ws2.cell(3 + i, 2).border = border_thin

        ws2.cell(3 + i, 3, f"{v:.3f}")
        ws2.cell(3 + i, 3).font = normal_font
        ws2.cell(3 + i, 3).alignment = center_align
        ws2.cell(3 + i, 3).border = border_thin

    # Adjust column widths
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()
    buf.close()

    return excel_bytes
